from openpyxl import load_workbook

from . import config, conversions
from . import typing as t

data_filepath = config.DATA_DIR / "data.xlsx"
workbook = load_workbook(filename=data_filepath, read_only=True)
worksheet = workbook.active
irows = worksheet.iter_rows(values_only=True)
header_row = next(irows)
# note that subsequent rows generated/iterated from `irows` are data rows

# get the number of the respective columns where each candidate's ranking is recorded
candidates: set[t.Candidate] = {"Lucky", "Somto", "Hosanna", "Tellah", "Chelsea"}
candidate_indexes: dict[int, t.Candidate] = {
    col_num: col_data
    for (col_num, col_data) in enumerate(header_row)
    if col_data in candidates
}

# extract the votes from each column
votes: list[t.Vote] = list(
    {
        candidate_indexes[col_num]: conversions.lookup_rank(col_data)
        for (col_num, col_data) in enumerate(data_row)
        if col_num in candidate_indexes
    }
    for data_row in irows
)
